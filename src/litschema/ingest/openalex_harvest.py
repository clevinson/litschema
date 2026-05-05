"""Phase 1a: Harvest structured metadata from OpenAlex API.

Queries OpenAlex for each DOI in the tracking spreadsheet. Extracts authors,
affiliations, ORCIDs, ROR IDs, abstracts, keywords. Saves raw JSON per paper.

Usage:
    uv run python -m litschema.ingest.openalex_harvest [--email EMAIL] [--data-dir DIR]
"""

from __future__ import annotations

import argparse
import json
import logging
import re
import time
from pathlib import Path

import requests

from ..config import load_config

logger = logging.getLogger(__name__)

_CFG = load_config()
# Project root is whatever litschema.yaml points at.
PROJECT_ROOT = _CFG.project_root
TRACKING_XLSX = _CFG.tracking_xlsx
DEFAULT_DATA_DIR = _CFG.openalex_dir

OPENALEX_API = "https://api.openalex.org/works"
# Rate limit: 10 req/s without polite pool, 100 req/s with mailto
RATE_LIMIT_DELAY = 0.1  # 100 req/s with polite pool


def doi_to_slug(doi: str) -> str:
    """Convert a DOI to a filesystem-safe slug."""
    return re.sub(r"[^a-zA-Z0-9]", "_", doi).strip("_").lower()


def reconstruct_abstract(inverted_index: dict | None) -> str | None:
    """Reconstruct abstract from OpenAlex inverted index format."""
    if not inverted_index:
        return None
    # inverted_index: {"word": [pos1, pos2, ...], ...}
    word_positions: list[tuple[int, str]] = []
    for word, positions in inverted_index.items():
        for pos in positions:
            word_positions.append((pos, word))
    word_positions.sort(key=lambda x: x[0])
    return " ".join(word for _, word in word_positions)


def load_dois_from_xlsx(xlsx_path: Path) -> list[dict]:
    """Load DOIs and metadata from tracking spreadsheet."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row, strict=False))
        doi = record.get("doi")
        if doi and str(doi).strip():
            doi = str(doi).strip()
            # Clean BOM characters
            doi = doi.lstrip("\ufeff")
            record["doi"] = doi
            rows.append(record)
    wb.close()
    return rows


def fetch_openalex(doi: str, email: str | None = None) -> dict | None:
    """Fetch a single work from OpenAlex by DOI."""
    params = {}
    if email:
        params["mailto"] = email

    url = f"{OPENALEX_API}/doi:{doi}"
    try:
        resp = requests.get(url, params=params, timeout=30)
        if resp.status_code == 200:
            return resp.json()
        elif resp.status_code == 404:
            logger.warning("Not found in OpenAlex: %s", doi)
            return None
        else:
            logger.error("OpenAlex error %d for %s: %s", resp.status_code, doi, resp.text[:200])
            return None
    except requests.RequestException as e:
        logger.error("Request failed for %s: %s", doi, e)
        return None


def extract_metadata(raw: dict) -> dict:
    """Extract structured metadata from OpenAlex response."""
    result = {
        "openalex_id": raw.get("id"),
        "doi": raw.get("doi", "").replace("https://doi.org/", ""),
        "title": raw.get("title"),
        "publication_year": raw.get("publication_year"),
        "publication_date": raw.get("publication_date"),
        "type": raw.get("type"),
        "open_access": raw.get("open_access", {}),
    }

    # Abstract
    result["abstract"] = reconstruct_abstract(raw.get("abstract_inverted_index"))

    # Primary location (journal info)
    primary = raw.get("primary_location") or {}
    source = primary.get("source") or {}
    result["journal"] = source.get("display_name")
    result["journal_issn"] = source.get("issn_l")
    result["publisher_from_source"] = source.get("host_organization_name")

    # Authors with affiliations
    authors = []
    for authorship in raw.get("authorships", []):
        author_data = authorship.get("author", {})
        institutions = []
        for inst in authorship.get("institutions", []):
            institutions.append(
                {
                    "name": inst.get("display_name"),
                    "ror": inst.get("ror"),
                    "country_code": inst.get("country_code"),
                    "type": inst.get("type"),
                }
            )
        # Also check raw_affiliation_strings for cases where institution wasn't matched
        raw_affiliations = authorship.get("raw_affiliation_strings", [])

        orcid = author_data.get("orcid")
        if orcid:
            orcid = orcid.replace("https://orcid.org/", "")

        authors.append(
            {
                "openalex_author_id": author_data.get("id"),
                "display_name": author_data.get("display_name"),
                "orcid": orcid,
                "author_position": authorship.get("author_position"),
                "is_corresponding": authorship.get("is_corresponding"),
                "institutions": institutions,
                "raw_affiliation_strings": raw_affiliations,
            }
        )
    result["authors"] = authors

    # Keywords and topics
    result["keywords"] = [kw.get("keyword") for kw in raw.get("keywords", []) if kw.get("keyword")]
    result["topics"] = [
        {"name": t.get("display_name"), "subfield": t.get("subfield", {}).get("display_name")}
        for t in raw.get("topics", [])
    ]

    # Citation info
    result["cited_by_count"] = raw.get("cited_by_count")
    result["referenced_works_count"] = raw.get("referenced_works_count")

    return result


def harvest(
    xlsx_path: Path = TRACKING_XLSX,
    data_dir: Path = DEFAULT_DATA_DIR,
    email: str | None = None,
    skip_existing: bool = True,
) -> dict:
    """Run the full OpenAlex harvest.

    Returns summary stats dict.
    """
    data_dir.mkdir(parents=True, exist_ok=True)

    rows = load_dois_from_xlsx(xlsx_path)
    logger.info("Loaded %d papers with DOIs from %s", len(rows), xlsx_path.name)

    stats = {"total": len(rows), "fetched": 0, "skipped": 0, "not_found": 0, "errors": 0}

    for i, row in enumerate(rows):
        doi = row["doi"]
        slug = doi_to_slug(doi)
        out_path = data_dir / f"{slug}.json"

        if skip_existing and out_path.exists():
            stats["skipped"] += 1
            continue

        raw = fetch_openalex(doi, email=email)
        if raw is None:
            stats["not_found"] += 1
            # Save a marker so we don't re-query
            out_path.write_text(json.dumps({"doi": doi, "error": "not_found"}))
        else:
            extracted = extract_metadata(raw)
            extracted["_source_doi"] = doi
            extracted["_tracking_author"] = row.get("author_short")
            extracted["_tracking_filename"] = row.get("filename")
            out_path.write_text(json.dumps(extracted, indent=2, ensure_ascii=False))
            stats["fetched"] += 1

        # Progress
        done = stats["fetched"] + stats["not_found"] + stats["skipped"]
        if (done % 50 == 0) or (i == len(rows) - 1):
            logger.info(
                "Progress: %d/%d (fetched=%d, skipped=%d, not_found=%d)",
                done,
                stats["total"],
                stats["fetched"],
                stats["skipped"],
                stats["not_found"],
            )

        time.sleep(RATE_LIMIT_DELAY)

    logger.info("Harvest complete: %s", stats)
    return stats


def main():
    parser = argparse.ArgumentParser(description="Harvest ERW paper metadata from OpenAlex")
    parser.add_argument("--email", help="Email for OpenAlex polite pool (recommended)")
    parser.add_argument("--data-dir", type=Path, default=DEFAULT_DATA_DIR)
    parser.add_argument("--xlsx", type=Path, default=TRACKING_XLSX)
    parser.add_argument("--no-skip", action="store_true", help="Re-fetch existing files")
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")

    stats = harvest(
        xlsx_path=args.xlsx,
        data_dir=args.data_dir,
        email=args.email,
        skip_existing=not args.no_skip,
    )
    print(json.dumps(stats, indent=2))


if __name__ == "__main__":
    main()
