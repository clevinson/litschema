"""Phase 5: Assemble all data into a LinkML-valid corpus.yaml.

Merges bibliographic fields (OpenAlex → CrossRef fallback), entity registries
(authors, institutions), and tracking spreadsheet metadata into a single
corpus.yaml that validates against the ERW articles schema.

Usage:
    uv run python -m litschema.ingest.assemble_corpus
"""

from __future__ import annotations

import argparse
import json
import logging
import re
from pathlib import Path

import yaml

from ..articles import article_files
from ..config import load_config

logger = logging.getLogger(__name__)

_CFG = load_config()
PROJECT_ROOT = _CFG.project_root
OPENALEX_DIR = _CFG.openalex_dir
CROSSREF_DIR = _CFG.crossref_dir
DATA_DIR = _CFG.data_dir
TRACKING_XLSX = _CFG.tracking_xlsx
CORPUS_OUT = _CFG.corpus_file

# Map CrossRef type to our DocumentTypeEnum
CROSSREF_TYPE_MAP = {
    "journal-article": "journal_article",
    "review-article": "review_article",
    "book-chapter": "book_chapter",
    "proceedings-article": "conference_paper",
    "posted-content": "preprint",
    "dissertation": "thesis",
    "report": "technical_report",
    "monograph": "book_chapter",
}

# Map OpenAlex type to our DocumentTypeEnum
OPENALEX_TYPE_MAP = {
    "article": "journal_article",
    "review": "review_article",
    "book-chapter": "book_chapter",
    "proceedings-article": "conference_paper",
    "preprint": "preprint",
    "dissertation": "thesis",
    "editorial": "letter",
    "letter": "letter",
    "erratum": "letter",
    "report": "technical_report",
    "paratext": None,  # Skip front matter etc.
    "dataset": "data_paper",
}


def load_llm_extraction(article_id: str, extractions_dir: Path | None = None) -> dict | None:
    """Load LLM extraction for an article if available."""
    if extractions_dir is None:
        files = article_files(_CFG, article_id)
        json_path = files.extraction_path()
        yaml_path = files.legacy_extraction.with_suffix(".yaml")
    else:
        # Explicit directory arguments preserve the historical flat-folder API.
        json_path = extractions_dir / f"{article_id}.json"
        yaml_path = extractions_dir / f"{article_id}.yaml"

    if json_path.exists():
        data = json.loads(json_path.read_text())
    elif yaml_path.exists():
        data = yaml.safe_load(yaml_path.read_text())
    else:
        return None
    if data and data.get("error"):
        return None
    return data


def merge_llm_extraction(article: dict, extractions_dir: Path | None = None) -> None:
    """Merge LLM extraction fields into an article dict in-place."""
    extraction = load_llm_extraction(article["id"], extractions_dir)
    if not extraction:
        return
    for field in (
        "study_types",
        "focus_areas",
        "document_type",
        "experimental_setups",
        "quantification_approaches",
        "modeling_details",
    ):
        value = extraction.get(field)
        if value:
            article[field] = value


def doi_to_slug(doi: str) -> str:
    return re.sub(r"[^a-zA-Z0-9]", "_", doi).strip("_").lower()


def make_article_id(author_short: str, year: int | None) -> str:
    """Generate article ID from tracking spreadsheet fields."""
    if not author_short:
        author_short = "unknown"
    # Extract first author last name
    author = author_short.split(" et al.")[0].split(",")[0].split(" and ")[0].strip()
    author = author.lower().replace(" ", "-")
    # Remove non-alphanumeric except hyphen
    author = re.sub(r"[^a-z0-9-]", "", author)
    year_str = str(year) if year else "nd"
    return f"{author}-{year_str}"


def load_tracking(xlsx_path: Path) -> dict[str, dict]:
    """Load tracking spreadsheet, keyed by DOI."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    by_doi: dict[str, dict] = {}
    no_doi: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row, strict=False))
        doi = record.get("doi")
        if doi and str(doi).strip():
            doi = str(doi).strip().lstrip("\ufeff")
            by_doi[doi] = record
        else:
            no_doi.append(record)
    wb.close()
    return by_doi, no_doi


def load_entity_registries(data_dir: Path) -> tuple[list[dict], list[dict]]:
    """Load institution and author registries."""
    inst_path = data_dir / "institutions.yaml"
    auth_path = data_dir / "authors.yaml"

    institutions = yaml.safe_load(inst_path.read_text()) if inst_path.exists() else []
    authors = yaml.safe_load(auth_path.read_text()) if auth_path.exists() else []

    return institutions, authors


def resolve_document_type(openalex_data: dict, crossref_data: dict | None) -> str | None:
    """Determine document_type from API data."""
    # Try OpenAlex type first
    oa_type = openalex_data.get("type")
    if oa_type and oa_type in OPENALEX_TYPE_MAP:
        mapped = OPENALEX_TYPE_MAP[oa_type]
        if mapped:
            return mapped

    # Fall back to CrossRef
    if crossref_data and not crossref_data.get("error"):
        cr_type = crossref_data.get("type")
        if cr_type and cr_type in CROSSREF_TYPE_MAP:
            return CROSSREF_TYPE_MAP[cr_type]

    return "journal_article"  # Default assumption for academic papers


def build_article(
    doi: str,
    tracking: dict,
    openalex_data: dict,
    crossref_data: dict | None,
    author_lookup: dict[str, str],  # openalex_author_id -> our author_id
    used_article_ids: set[str],
) -> dict:
    """Build a single article record."""
    # Article ID
    base_id = make_article_id(
        tracking.get("author_short", ""),
        openalex_data.get("publication_year") or tracking.get("year"),
    )
    article_id = base_id
    suffix = 2
    while article_id in used_article_ids:
        article_id = f"{base_id}-{suffix}"
        suffix += 1
    used_article_ids.add(article_id)

    article = {"id": article_id}

    # DOI - validate format before including
    clean_doi = doi.replace("https://doi.org/", "").replace("http://doi.org/", "").strip()
    if re.match(r"^10\.\d{4,9}/\S+$", clean_doi):
        article["doi"] = clean_doi
    else:
        # Invalid DOI (e.g., ISSNs, malformed entries) - store in notes
        article["notes"] = f"Invalid DOI in source: {clean_doi}"
        logger.warning("Skipping invalid DOI: %s", clean_doi)

    # Title - prefer OpenAlex
    article["title"] = openalex_data.get("title") or tracking.get("title", "")

    # Year
    year = openalex_data.get("publication_year") or tracking.get("year")
    if year:
        article["year"] = int(year)

    # Filename from tracking
    filename = tracking.get("filename")
    if filename:
        article["filename"] = str(filename)

    # Authors - cross-reference by OpenAlex author ID
    author_ids = []
    for auth in openalex_data.get("authors", []):
        oa_author_id = auth.get("openalex_author_id")
        if oa_author_id and oa_author_id in author_lookup:
            author_ids.append(author_lookup[oa_author_id])
    if author_ids:
        article["author_ids"] = author_ids

    # Document type
    doc_type = resolve_document_type(openalex_data, crossref_data)
    if doc_type:
        article["document_type"] = doc_type

    # Publisher - from tracking (already curated)
    publisher = tracking.get("publisher")
    if publisher:
        article["publisher"] = publisher

    # Journal
    journal = openalex_data.get("journal")
    if journal:
        article["journal"] = journal

    # Abstract
    abstract = openalex_data.get("abstract")
    if not abstract and crossref_data and not crossref_data.get("error"):
        abstract = crossref_data.get("abstract")
    if abstract:
        article["abstract"] = abstract

    # Keywords
    keywords = openalex_data.get("keywords", [])
    if keywords:
        article["keywords"] = keywords

    # Peer reviewed - assume true for journal articles
    if doc_type in ("journal_article", "review_article"):
        article["peer_reviewed"] = True
    elif doc_type in ("preprint", "thesis", "white_paper", "blog_post"):
        article["peer_reviewed"] = False

    # Notes from tracking
    notes = tracking.get("notes")
    if notes:
        article["notes"] = str(notes)

    # Merge LLM extraction (study_types, experimental_setups, etc.)
    merge_llm_extraction(article)

    return article


def build_no_doi_article(tracking: dict, used_article_ids: set[str]) -> dict:
    """Build article record for papers without DOIs (minimal metadata)."""
    base_id = make_article_id(
        tracking.get("author_short", ""),
        tracking.get("year"),
    )
    article_id = base_id
    suffix = 2
    while article_id in used_article_ids:
        article_id = f"{base_id}-{suffix}"
        suffix += 1
    used_article_ids.add(article_id)

    article = {"id": article_id}
    article["title"] = tracking.get("title", "")
    year = tracking.get("year")
    if year:
        article["year"] = int(year)

    filename = tracking.get("filename")
    if filename:
        article["filename"] = str(filename)

    publisher = tracking.get("publisher")
    if publisher:
        article["publisher"] = publisher

    notes = tracking.get("notes")
    if notes:
        article["notes"] = str(notes)

    # Merge LLM extraction (study_types, experimental_setups, etc.)
    merge_llm_extraction(article)

    return article


def assemble(
    openalex_dir: Path = OPENALEX_DIR,
    crossref_dir: Path = CROSSREF_DIR,
    data_dir: Path = DATA_DIR,
    tracking_xlsx: Path = TRACKING_XLSX,
    output: Path = CORPUS_OUT,
) -> dict:
    """Assemble the full corpus. Returns summary stats."""
    # Load registries
    institutions, authors = load_entity_registries(data_dir)

    # Build author lookup: OpenAlex ID -> our author ID
    author_lookup: dict[str, str] = {}
    for auth in authors:
        oa_id = auth.get("_openalex_id")
        if oa_id:
            author_lookup[oa_id] = auth["id"]

    # Clean internal fields from authors before output
    clean_authors = []
    for auth in authors:
        clean = {k: v for k, v in auth.items() if not k.startswith("_") and v is not None}
        clean_authors.append(clean)

    # Load tracking data
    by_doi, no_doi_papers = load_tracking(tracking_xlsx)

    # Load OpenAlex and CrossRef data
    openalex_by_doi: dict[str, dict] = {}
    for json_file in sorted(openalex_dir.glob("*.json")):
        data = json.loads(json_file.read_text())
        doi = data.get("_source_doi") or data.get("doi")
        if doi:
            openalex_by_doi[doi] = data

    crossref_by_doi: dict[str, dict] = {}
    if crossref_dir.exists():
        for json_file in sorted(crossref_dir.glob("*.json")):
            data = json.loads(json_file.read_text())
            doi = data.get("doi")
            if doi:
                crossref_by_doi[doi] = data

    # Build articles
    articles = []
    used_article_ids: set[str] = set()
    stats = {"with_doi": 0, "without_doi": 0, "with_authors": 0, "with_abstract": 0}

    for doi, tracking in sorted(by_doi.items()):
        openalex_data = openalex_by_doi.get(doi, {})
        crossref_data = crossref_by_doi.get(doi)

        if openalex_data.get("error"):
            openalex_data = {}  # Treat as empty

        article = build_article(
            doi,
            tracking,
            openalex_data,
            crossref_data,
            author_lookup,
            used_article_ids,
        )
        articles.append(article)
        stats["with_doi"] += 1
        if article.get("author_ids"):
            stats["with_authors"] += 1
        if article.get("abstract"):
            stats["with_abstract"] += 1

    # No-DOI papers
    for tracking in no_doi_papers:
        article = build_no_doi_article(tracking, used_article_ids)
        articles.append(article)
        stats["without_doi"] += 1

    # Assemble corpus
    corpus = {
        "articles": articles,
        "authors": clean_authors,
        "institutions": institutions,
    }

    # Write YAML
    with open(output, "w") as f:
        f.write("# ERW Research Corpus\n")
        f.write("# Generated by litschema.ingest.assemble_corpus\n")
        f.write(
            f"# {len(articles)} articles, {len(clean_authors)} authors, {len(institutions)} institutions\n\n"
        )
        yaml.dump(
            corpus, f, default_flow_style=False, allow_unicode=True, sort_keys=False, width=120
        )

    stats["total_articles"] = len(articles)
    stats["total_authors"] = len(clean_authors)
    stats["total_institutions"] = len(institutions)
    logger.info("Corpus assembled: %s", stats)
    logger.info("Written to %s", output)
    return stats


def main():
    parser = argparse.ArgumentParser(description="Assemble LinkML corpus from harvested data")
    parser.add_argument("--openalex-dir", type=Path, default=OPENALEX_DIR)
    parser.add_argument("--crossref-dir", type=Path, default=CROSSREF_DIR)
    parser.add_argument("--data-dir", type=Path, default=DATA_DIR)
    parser.add_argument("--tracking", type=Path, default=TRACKING_XLSX)
    parser.add_argument("--output", type=Path, default=CORPUS_OUT)
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    stats = assemble(
        openalex_dir=args.openalex_dir,
        crossref_dir=args.crossref_dir,
        data_dir=args.data_dir,
        tracking_xlsx=args.tracking,
        output=args.output,
    )
    print(json.dumps(stats, indent=2))


if __name__ == "__main__":
    main()
